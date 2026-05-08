"""
app.py
─────────────────────────────────────────────────────────────────────────────
Student Data Management System – Flask application entry point.

Routes
──────
GET  /                              → Render single-page frontend
GET  /api/students                  → Retrieve all students
POST /api/students                  → Add a new student
PUT  /api/students/<student_id>     → Update an existing student
DELETE /api/students/<student_id>   → Delete a student
GET  /api/students/search           → Search students (linear/binary/sequential)
GET  /api/students/sort             → Sort students (bubble/insertion/selection/merge/shell)

Error handling: every route wraps its business logic in try/except blocks
and returns structured JSON error responses instead of raw exceptions.
─────────────────────────────────────────────────────────────────────────────
"""

import os
import tempfile
import time

from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    url_for,
    send_file,
)
from flask_login import (
    LoginManager,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, LongTable
from reportlab.lib.styles import getSampleStyleSheet
import io

from models.firestore_student_manager import FirestoreStudentManager
from models.firestore_user_manager import FirestoreUserManager
from utils.validators import StudentValidator
from utils.search import SearchAlgorithms
from utils.sort import SortAlgorithms


# ─────────────────────────── App setup ────────────────────────────────────

app = Flask(__name__)
app.secret_key = "replace-with-a-secure-key"


def _resolve_data_file(name: str) -> str:
    preferred_dir = os.environ.get("DATA_DIR") or os.path.join(app.root_path, "data")
    fallback_dir = os.path.join(tempfile.gettempdir(), "data_mahasiswa")

    for data_dir in (preferred_dir, fallback_dir):
        try:
            os.makedirs(data_dir, exist_ok=True)
            test_file = os.path.join(data_dir, ".permtest")
            with open(test_file, "a", encoding="utf-8"):
                pass
            os.remove(test_file)
            return os.path.join(data_dir, name)
        except OSError:
            continue

    raise RuntimeError("Tidak ada direktori data yang dapat ditulis. Pastikan aplikasi memiliki izin penulisan.")

manager = FirestoreStudentManager(seed_file=_resolve_data_file("students.json"))
manager.backfill_missing_semesters(3, 7)
manager.backfill_missing_profiles()
user_manager = FirestoreUserManager(seed_file=_resolve_data_file("users.json"))

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Silakan login terlebih dahulu."


@login_manager.user_loader
def load_user(user_id: str):
    return user_manager.get_user_by_id(user_id)


# ─────────────────────────── Helper ───────────────────────────────────────

def _error(messages, status: int = 400):
    """Return a structured JSON error response."""
    if isinstance(messages, str):
        messages = [messages]
    return jsonify({"success": False, "errors": messages}), status


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Username dan password harus diisi.", "danger")
            return render_template("login.html")

        user = user_manager.verify_user(username, password)
        if user:
            login_user(user)
            next_page = request.args.get("next")
            return redirect(next_page or url_for("index"))

        flash("Username atau password salah.", "danger")

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not username or not email or not password or not confirm_password:
            flash("Semua kolom harus diisi.", "danger")
            return render_template("register.html")

        if password != confirm_password:
            flash("Password dan konfirmasi tidak cocok.", "danger")
            return render_template("register.html")

        try:
            user_manager.create_user(username=username, email=email, password=password, role="admin")
            flash("Registrasi berhasil. Silakan login.", "success")
            return redirect(url_for("login"))
        except ValueError as exc:
            flash(str(exc), "danger")
        except Exception as exc:
            flash("Terjadi kesalahan saat mendaftarkan akun. Silakan coba lagi nanti.", "danger")
            app.logger.error("Registrasi gagal: %s", exc)

    return render_template("register.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not username or not email or not new_password or not confirm_password:
            flash("Semua kolom harus diisi.", "danger")
            return render_template("forgot_password.html")

        if new_password != confirm_password:
            flash("Password dan konfirmasi password tidak cocok.", "danger")
            return render_template("forgot_password.html")

        success = user_manager.reset_password(username, email, new_password)
        if success:
            flash("Password berhasil diubah. Silakan login dengan password baru Anda.", "success")
            return redirect(url_for("login"))
        else:
            flash("Username atau email tidak sesuai.", "danger")

    return render_template("forgot_password.html")


@app.route("/forgot-password/verify", methods=["POST"])
def verify_forgot_password():
    if current_user.is_authenticated:
        return jsonify({"success": False, "message": "Anda sudah login."}), 403

    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip().lower()
    email = data.get("email", "").strip().lower()

    if not username or not email:
        return jsonify({"success": False, "message": "Username dan email harus diisi."}), 400

    user = user_manager.get_user_by_username(username)
    if user is None or user.email.lower() != email:
        return jsonify({"success": False, "message": "Username atau email tidak sesuai."}), 404

    return jsonify({"success": True, "message": "Data terverifikasi. Silakan masukkan password baru."})


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/admin/users")
@login_required
def admin_users():
    if not current_user.is_super_admin:
        abort(403)

    users = user_manager.get_all_users()
    return render_template("admin.html", users=users)


@app.route("/admin/users/<user_id>/promote", methods=["POST"])
@login_required
def promote_user(user_id: str):
    if not current_user.is_super_admin:
        abort(403)

    user = user_manager.get_user_by_id(user_id)
    if user is None:
        abort(404)

    if user.is_admin:
        flash(f"User '{user.username}' sudah administrator.", "info")
    else:
        user_manager.set_user_role(user_id, "admin")
        flash(f"User '{user.username}' berhasil dijadikan administrator.", "success")

    return redirect(url_for("admin_users"))


@app.route("/admin/users/<user_id>/demote", methods=["POST"])
@login_required
def demote_user(user_id: str):
    if not current_user.is_super_admin:
        abort(403)

    user = user_manager.get_user_by_id(user_id)
    if user is None:
        abort(404)

    if user.get_id() == current_user.get_id():
        flash("Anda tidak dapat mengubah role diri sendiri.", "warning")
    elif not user.is_admin:
        flash(f"User '{user.username}' bukan administrator.", "info")
    else:
        user_manager.set_user_role(user_id, "user")
        flash(f"User '{user.username}' berhasil dikembalikan menjadi user biasa.", "success")

    return redirect(url_for("admin_users"))


@app.route("/admin/users/<user_id>/delete", methods=["POST"])
@login_required
def delete_user(user_id: str):
    if not current_user.is_super_admin:
        abort(403)

    if user_id == current_user.get_id():
        flash("Anda tidak dapat menghapus diri sendiri.", "warning")
        return redirect(url_for("admin_users"))

    user = user_manager.get_user_by_id(user_id)
    if user is None:
        abort(404)

    deleted = user_manager.delete_user(user_id)
    if deleted:
        flash(f"User '{user.username}' berhasil dihapus.", "success")
    else:
        flash(f"User '{user.username}' tidak ditemukan.", "warning")

    return redirect(url_for("admin_users"))


# ─────────────────────────── Page Route ───────────────────────────────────

@app.route("/")
@login_required
def index():
    """Serve the single-page frontend."""
    return render_template("index.html")


# ─────────────────────────── Students – CRUD ──────────────────────────────

@app.route("/api/students", methods=["GET"])
@login_required
def get_students():
    """
    Return all students as a JSON array.

    Time Complexity: O(n) – serialises every student.
    """
    try:
        manager.reload()
        return jsonify({"success": True, "data": manager.get_all_dicts()})
    except Exception as exc:
        return _error(str(exc), 500)


@app.route("/api/students", methods=["POST"])
@login_required
def add_student():
    """
    Create and persist a new student.

    Validates the request body with regex-based rules before delegating to
    the manager.  Returns 201 on success, 400 on validation failure.

    Time Complexity: O(n) – duplicate-ID check scans the list.
    """
    if not current_user.is_admin:
        abort(403)

    try:
        data = request.get_json(force=True, silent=True)
        if data is None:
            return _error("Request body must be valid JSON.")

        errors = StudentValidator.validate(data, is_update=False)
        if errors:
            return _error(errors, 400)

        student = manager.add_student(data)
        return jsonify({"success": True, "data": student.to_dict()}), 201

    except ValueError as exc:
        return _error(str(exc), 400)
    except IOError as exc:
        return _error(f"Storage error: {exc}", 500)
    except Exception as exc:
        return _error("An unexpected error occurred.", 500)


@app.route("/api/students/<student_id>", methods=["PUT"])
@login_required
def update_student(student_id: str):
    """
    Partially update an existing student.

    Only fields present in the request body are updated (PATCH semantics via
    PUT). student_id comes from the URL and is never re-validated here.

    Time Complexity: O(n) – ID lookup.
    """
    if not current_user.is_admin:
        abort(403)

    try:
        data = request.get_json(force=True, silent=True)
        if data is None:
            return _error("Request body must be valid JSON.")

        errors = StudentValidator.validate(data, is_update=True)
        if errors:
            return _error(errors, 400)

        student = manager.update_student(student_id, data)
        if student is None:
            return _error(f"Student '{student_id}' not found.", 404)

        return jsonify({"success": True, "data": student.to_dict()})

    except ValueError as exc:
        return _error(str(exc), 400)
    except IOError as exc:
        return _error(f"Storage error: {exc}", 500)
    except Exception as exc:
        return _error("An unexpected error occurred.", 500)


@app.route("/api/students/<student_id>", methods=["DELETE"])
@login_required
def delete_student(student_id: str):
    """
    Delete a student by ID.

    Time Complexity: O(n) – list rebuild after removal.
    """
    if not current_user.is_admin:
        abort(403)

    try:
        deleted = manager.delete_student(student_id)
        if not deleted:
            return _error(f"Student '{student_id}' not found.", 404)

        return jsonify({"success": True, "message": "Student deleted successfully."})

    except IOError as exc:
        return _error(f"Storage error: {exc}", 500)
    except Exception as exc:
        return _error("An unexpected error occurred.", 500)


# ─────────────────────────── Search ───────────────────────────────────────

@app.route("/api/students/search", methods=["GET"])
@login_required
def search_students():
    """
    Search students by a field value using a chosen algorithm.

    Query parameters
    ────────────────
    query     : str  – search term (required)
    field     : str  – name | student_id | email | major | phone  (default: name)
    algorithm : str  – linear | binary | sequential                (default: linear)

    Time complexities
    ─────────────────
    linear     : O(n)
    sequential : O(n) worst, O(1) best
    binary     : O(log n) search + O(k) collection
                 ⚠ requires sorted data – the list is sorted by merge sort first
    """
    try:
        manager.reload()
        query     = request.args.get("query", "").strip()
        field     = request.args.get("field", "name").lower()
        algorithm = request.args.get("algorithm", "linear").lower()

        if not query:
            return _error("Query parameter 'query' cannot be empty.")

        valid_algorithms = {"linear", "binary", "sequential"}
        if algorithm not in valid_algorithms:
            return _error(f"Unknown algorithm '{algorithm}'. "
                          f"Choose from: {', '.join(valid_algorithms)}.")

        students = manager.get_all()

        # ── Measure execution time ──────────────────────────────────────
        start = time.perf_counter()

        if algorithm == "binary":
            # Binary search requires sorted input – use merge sort (O(n log n))
            sorted_students = SortAlgorithms.merge_sort(students, key=field)
            results = SearchAlgorithms.binary_search(sorted_students, query, field)
            complexity = {"best": "O(1)", "avg": "O(log n)", "worst": "O(log n)"}

        elif algorithm == "sequential":
            results = SearchAlgorithms.sequential_search(students, query, field)
            complexity = {"best": "O(1)", "avg": "O(n)", "worst": "O(n)"}

        else:   # linear (default)
            results = SearchAlgorithms.linear_search(students, query, field)
            complexity = {"best": "O(1)", "avg": "O(n)", "worst": "O(n)"}

        elapsed_ms = round((time.perf_counter() - start) * 1_000, 4)

        return jsonify({
            "success":    True,
            "data":       [s.to_dict() for s in results],
            "algorithm":  algorithm,
            "complexity": complexity,
            "elapsed_ms": elapsed_ms,
            "count":      len(results),
        })

    except Exception as exc:
        return _error(str(exc), 500)


# ─────────────────────────── Sort ─────────────────────────────────────────

@app.route("/api/students/sort", methods=["GET"])
@login_required
def sort_students():
    """
    Sort students by a chosen field using a chosen algorithm.

    Query parameters
    ────────────────
    algorithm : str  – bubble | insertion | selection | merge | shell
    field     : str  – name | student_id | gpa | age | major | email
    order     : str  – asc | desc  (default: asc)

    Time / space complexities are returned in the response for reference.
    """
    try:
        manager.reload()
        algorithm = request.args.get("algorithm", "bubble").lower()
        field     = request.args.get("field", "name").lower()
        order     = request.args.get("order", "asc").lower()

        # Map algorithm name → (sort function, time complexity, space complexity)
        algo_map = {
            "bubble":    (SortAlgorithms.bubble_sort,    "O(n²)",       "O(1)"),
            "insertion": (SortAlgorithms.insertion_sort, "O(n²)",       "O(1)"),
            "selection": (SortAlgorithms.selection_sort, "O(n²)",       "O(1)"),
            "merge":     (SortAlgorithms.merge_sort,     "O(n log n)",  "O(n)"),
            "shell":     (SortAlgorithms.shell_sort,     "O(n log²n)",  "O(1)"),
        }

        if algorithm not in algo_map:
            return _error(
                f"Unknown algorithm '{algorithm}'. "
                f"Choose from: {', '.join(algo_map)}."
            )

        sort_fn, time_complexity, space_complexity = algo_map[algorithm]
        students = manager.get_all()

        start = time.perf_counter()
        sorted_students = sort_fn(students, key=field)
        elapsed_ms = round((time.perf_counter() - start) * 1_000, 4)

        if order == "desc":
            sorted_students = sorted_students[::-1]

        return jsonify({
            "success":          True,
            "data":             [s.to_dict() for s in sorted_students],
            "algorithm":        algorithm,
            "field":            field,
            "order":            order,
            "time_complexity":  time_complexity,
            "space_complexity": space_complexity,
            "elapsed_ms":       elapsed_ms,
        })

    except Exception as exc:
        return _error(str(exc), 500)


# ─────────────────────────── Export ───────────────────────────────────────

@app.route("/api/students/export/excel", methods=["GET"])
@login_required
def export_students_excel():
    """
    Export all students to an Excel file.
    """
    try:
        manager.reload()
        students = manager.get_all()

        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        # Header
        headers = ["NIM", "Nama", "Tanggal Lahir", "Email", "Telepon", "Jenjang Pendidikan", "Jurusan", "IPK", "Nilai", "Semester", "Dibuat Pada"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Data
        for row_num, student in enumerate(students, 2):
            ws.cell(row=row_num, column=1, value=student.student_id)
            ws.cell(row=row_num, column=2, value=student.name)
            ws.cell(row=row_num, column=3, value=student.birth_date or "-")
            ws.cell(row=row_num, column=4, value=student.email)
            ws.cell(row=row_num, column=5, value=student.phone)
            ws.cell(row=row_num, column=6, value=student.education_level or "-")
            ws.cell(row=row_num, column=7, value=student.major)
            ws.cell(row=row_num, column=8, value=student.gpa)
            ws.cell(row=row_num, column=9, value=student.get_grade_letter())
            ws.cell(row=row_num, column=10, value=student.semester if student.semester is not None else "-")
            ws.cell(row=row_num, column=11, value=student.created_at)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="students.xlsx"
        )

    except Exception as exc:
        return _error(str(exc), 500)


@app.route("/api/students/export/pdf", methods=["GET"])
@login_required
def export_students_pdf():
    """
    Export all students to a PDF file.
    """
    try:
        manager.reload()
        students = manager.get_all()

        # Create PDF with landscape orientation
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter))
        elements = []

        # Title
        styles = getSampleStyleSheet()
        title = Paragraph("Ekspor Data Mahasiswa", styles['Title'])
        elements.append(title)

        # Data
        data = [["NIM", "Nama", "Tanggal Lahir", "Email", "Telepon", "Jenjang", "Jurusan", "IPK", "Nilai", "Semester"]]
        for student in students:
            data.append([
                student.student_id,
                student.name,
                student.birth_date or "-",
                student.email,
                student.phone,
                student.education_level or "-",
                student.major,
                f"{student.gpa:.2f}",
                student.get_grade_letter(),
                str(student.semester) if student.semester is not None else "-",
            ])

        # Define column widths (in points, 72 points = 1 inch)
        # Landscape letter is 11" x 8.5", so we have ~792 points width
        col_widths = [80, 100, 70, 135, 80, 55, 80, 45, 45, 55]

        # Create table with column widths
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),  # Smaller font
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),  # Smaller font for data
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('WORDWRAP', (0, 0), (-1, -1), True),  # Allow word wrapping
        ]))

        elements.append(table)
        doc.build(elements)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name="students.pdf"
        )

    except Exception as exc:
        return _error(str(exc), 500)


# ─────────────────────────── Entry point ──────────────────────────────────

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
